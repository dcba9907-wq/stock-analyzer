import json
import os
import subprocess
import sys
import threading
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from calculator import calculate_srim
from scorer import compute_scores
from universe_snapshot import list_snapshots, load_snapshot, compare_universes
from dart_api import (
    get_consensus_target_price, get_current_price, get_financial_data,
    get_quarterly_income, get_share_count, search_corp,
)

st.set_page_config(
    page_title="S-RIM 적정주가 계산기",
    page_icon="📈",
    layout="wide",
)

HISTORY_FILE = Path(__file__).parent / "history.json"
RESULTS_FILE = Path(__file__).parent / "scan_results.json"


# ─── Helpers ──────────────────────────────────────────────────────────────────

def load_history() -> list:
    if HISTORY_FILE.exists():
        return json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
    return []


def save_history(entry: dict):
    history = load_history()
    history.append(entry)
    HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")


def load_scan_results() -> list:
    if RESULTS_FILE.exists():
        return json.loads(RESULTS_FILE.read_text(encoding="utf-8"))
    return []


def fmt(n, dec: int = 0) -> str:
    if n is None:
        return "-"
    return f"{n:,.{dec}f}"


def analyze(name: str, price: int, year: str, req_ret: float, w: float) -> dict:
    corp_code, stock_code, matched = search_corp(name)
    if not corp_code:
        return {"error": "기업을 찾을 수 없습니다", "input_name": name}

    auto_price = False
    if price == 0 and stock_code:
        fetched = get_current_price(stock_code)
        if fetched:
            price = fetched
            auto_price = True

    target_price = get_consensus_target_price(stock_code) if stock_code else None

    fin = get_financial_data(corp_code, year)
    if not fin:
        return {"error": f"{year}년 재무데이터가 없습니다", "input_name": name}

    shares = get_share_count(corp_code, year)
    if not shares:
        return {"error": "주식총수 데이터를 가져올 수 없습니다", "input_name": name}

    calc = calculate_srim(fin["equity"], fin["net_income"], shares, req_ret, w)

    low_roe = calc["low_roe"]
    ratio_basic = price / calc["fair_basic"] if (not low_roe and calc["fair_basic"] != 0) else None
    ratio_w = price / calc["fair_w"] if (not low_roe and calc["fair_w"] != 0) else None

    return {
        "company": matched,
        "input_name": name,
        "corp_code": corp_code,
        "stock_code": stock_code,
        "current_price": price,
        "auto_price": auto_price,
        "equity": fin["equity"],
        "net_income": fin["net_income"],
        "shares": shares,
        "roe": calc["roe"],
        "low_roe": low_roe,
        "fair_basic": calc["fair_basic"],
        "fair_w": calc["fair_w"],
        "ratio_basic": ratio_basic,
        "ratio_w": ratio_w,
        "recommend_basic": (ratio_basic < 1) if ratio_basic is not None else False,
        "recommend_w": (ratio_w < 1) if ratio_w is not None else False,
        "target_price": target_price,
        "req_ret": req_ret,
    }


def render_card(result: dict, col, year: str):
    with col:
        if "error" in result:
            st.error(f"**{result.get('input_name', '')}**: {result['error']}")
            return

        low_roe = result.get("low_roe", False)
        rec = result["recommend_basic"]
        if low_roe:
            bg, border = "#fff8e1", "#f0ad4e"
        elif rec:
            bg, border = "#d4edda", "#28a745"
        else:
            bg, border = "#fde8e8", "#dc3545"

        eq_bil = result["equity"] / 1e8
        ni_bil = result["net_income"] / 1e8

        def row(label, value):
            return (
                f'<tr>'
                f'<td style="padding:4px 8px;color:#444;font-size:0.88em;">{label}</td>'
                f'<td style="padding:4px 8px;text-align:right;font-weight:600;font-size:0.92em;">{value}</td>'
                f'</tr>'
            )

        def divider():
            return '<tr><td colspan="2"><hr style="margin:4px 0;border:none;border-top:1px solid #ccc;"></td></tr>'

        warning_label = "⚠️ ROE < 요구수익률 (저수익 기업)"
        price_label = f"{fmt(result['current_price'])} 원"
        if result.get("auto_price"):
            price_label += " (자동조회)"

        if low_roe:
            fair_basic_label = f"{fmt(result['fair_basic'])} 원 (순자산가치)"
            fair_w_label = f"{fmt(result['fair_w'])} 원 (순자산가치)"
            ratio_basic_label = "-"
            ratio_w_label = "-"
            judge_basic = warning_label
            judge_w = warning_label
        else:
            fair_basic_label = f"{fmt(result['fair_basic'])} 원"
            fair_w_label = f"{fmt(result['fair_w'])} 원"
            ratio_basic_label = fmt(result["ratio_basic"], 2)
            ratio_w_label = fmt(result["ratio_w"], 2)
            judge_basic = "✅ 추천" if result["recommend_basic"] else "❌ 미추천"
            judge_w = "✅ 추천" if result["recommend_w"] else "❌ 미추천"

        tp = result.get("target_price")
        tp_label = f"{fmt(tp)} 원" if tp else "정보 없음"

        rows_html = (
            row("자기자본", f"{fmt(eq_bil)} 억원")
            + row("당기순이익", f"{fmt(ni_bil)} 억원")
            + row("ROE", f"{fmt(result['roe'], 2)} %")
            + row("발행주식수", f"{fmt(result['shares'])} 주")
            + divider()
            + row("적정주가 (기본)", fair_basic_label)
            + row("적정주가 (보수적)", fair_w_label)
            + row("평균 목표주가", f"{tp_label} <small style='color:#888;font-size:0.8em;'>(FnGuide)</small>")
            + row("현재종가", price_label)
            + divider()
            + row("종가/적정주가 (기본)", ratio_basic_label)
            + row("종가/적정주가 (보수적)", ratio_w_label)
            + divider()
            + row("판단 (기본)", judge_basic)
            + row("판단 (보수적)", judge_w)
        )

        st.markdown(
            f"""
            <div style="background:{bg};border:2px solid {border};border-radius:12px;
                        padding:18px 20px;margin-bottom:8px;">
              <h3 style="margin:0 0 12px 0;color:#222;">{result['company']}</h3>
              <table style="width:100%;border-collapse:collapse;">
                {rows_html}
              </table>
            </div>
            """,
            unsafe_allow_html=True,
        )

        corp_code = result.get("corp_code", "")
        equity = result["equity"]
        start_year = datetime.now().year
        qkey = f"quarterly_{corp_code}"
        with st.expander("📊 분기실적 보기 (참고용)"):
            st.caption("※ 참고용 — 적정주가 계산에 미반영")
            if qkey not in st.session_state:
                if st.button("분기 데이터 조회", key=f"qbtn_{corp_code}"):
                    with st.spinner("분기 데이터 조회 중..."):
                        st.session_state[qkey] = get_quarterly_income(corp_code, start_year)
            if qkey in st.session_state:
                quarters = st.session_state[qkey]
                if quarters:
                    rows = []
                    for q in quarters:
                        inc_label = f"{q['income']/1e8:,.0f}" if q["income"] is not None else "-"
                        if q["ttm"] is not None:
                            ttm_bil = f"{q['ttm']/1e8:,.0f}"
                            ttm_roe = f"{q['ttm'] / equity * 100:.2f} %" if equity else "-"
                        else:
                            ttm_bil = "-"
                            ttm_roe = "-"
                        rows.append({
                            "분기": f"{q['year']}년 {q['label']}",
                            "당기순이익 (억원)": inc_label,
                            "TTM 순이익 (억원)": ttm_bil,
                            "TTM ROE": ttm_roe,
                        })
                    st.dataframe(pd.DataFrame(rows), hide_index=True, use_container_width=True)
                    st.caption("TTM = 직전연도 연간 − 직전연도 동기누적 + 올해 누적")
                else:
                    st.info("분기 데이터가 없습니다.")


def make_excel(results: list, year: str, w: float) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S-RIM 분석"

    headers = [
        "기업명", "기준연도", "요구수익률(%)", "실적저하율(w)",
        "자기자본(억원)", "당기순이익(억원)", "ROE(%)", "발행주식수",
        "적정주가_기본(원)", "적정주가_보수적(원)", "현재종가(원)",
        "비율_기본", "비율_보수적", "판단_기본", "판단_보수적",
    ]

    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True)

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for r, res in enumerate(results, 2):
        if "error" in res:
            ws.cell(r, 1, res.get("input_name", ""))
            ws.cell(r, 2, "오류: " + res["error"])
            continue

        vals = [
            res["company"], year, res.get("req_ret", ""), w,
            round(res["equity"] / 1e8, 1),
            round(res["net_income"] / 1e8, 1),
            round(res["roe"], 2),
            res["shares"],
            round(res["fair_basic"]),
            round(res["fair_w"]),
            res["current_price"],
            round(res["ratio_basic"], 2) if res["ratio_basic"] is not None else "",
            round(res["ratio_w"], 2) if res["ratio_w"] is not None else "",
            "추천" if res["recommend_basic"] else "미추천",
            "추천" if res["recommend_w"] else "미추천",
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(r, c, v)

        color = "C6EFCE" if res["recommend_basic"] else "FFC7CE"
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).fill = fill

    for col in ws.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 4, 30)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def make_scan_excel(df: pd.DataFrame) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "전체스캔"

    headers = list(df.columns)
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for r, (_, row_data) in enumerate(df.iterrows(), 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(r, c, val)

    for col in ws.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 4, 28)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


_cy = datetime.now().year

# ─── Sidebar ──────────────────────────────────────────────────────────────────
# 조건부 렌더링 없이 모든 위젯을 항상 표시 — 탭 전환 시 화면이 바뀌는 문제 방지

with st.sidebar:
    st.header("⚙️ 계산 설정")

    # ── 개별분석 설정 ─────────────────────────────────────────────────────────
    st.subheader("🔍 개별분석")
    w = st.slider(
        "실적저하율 w",
        min_value=0.0, max_value=0.5, value=0.2, step=0.05, format="%.2f",
        key="sidebar_w",
    )
    year = st.selectbox(
        "기준연도",
        [str(_cy - i) for i in range(3)],
        key="sidebar_year",
    )

    st.divider()

    # ── 전체스캔 필터 ─────────────────────────────────────────────────────────
    st.subheader("📡 전체스캔 필터")
    st.number_input(
        "요구수익률 (%)",
        min_value=1.0, max_value=30.0, value=5.0, step=0.1, format="%.1f",
        key="scan_req_ret",
        help="시가총액 5,000억 이상 우량기업 기준 — kisrating AA- 등급 금리 참고",
    )
    st.slider(
        "종가/적정주가 상한",
        min_value=0.3, max_value=1.0, value=0.8, step=0.05,
        key="scan_ratio_max",
    )
    st.number_input(
        "시가총액 하한 (억원)",
        min_value=0, value=5000, step=500,
        key="scan_min_cap",
    )
    st.slider(
        "실적저하율 w (스캔용)",
        min_value=0.0, max_value=0.5, value=0.2, step=0.05, format="%.2f",
        key="scan_w",
    )

    st.divider()

    # ── 전체스캔 정렬 ─────────────────────────────────────────────────────────
    st.subheader("↕️ 정렬 기준")
    st.radio(
        "정렬 기준",
        ["종합점수순", "저평가순(비율)"],
        key="scan_sort_by",
        label_visibility="collapsed",
    )

    st.divider()
    st.link_button(
        "📊 한국신용평가 등급별 금리",
        "https://www.kisrating.com/ratingsStatistics/statics_spread.do",
        use_container_width=True,
    )


# ─── Main ─────────────────────────────────────────────────────────────────────

st.title("📈 S-RIM 적정주가 계산기")

if not os.getenv("DART_API_KEY"):
    st.error("⚠️ DART_API_KEY가 설정되지 않았습니다. .env 파일에 API 키를 추가하세요.")
    st.stop()

tab_individual, tab_scan, tab_universe = st.tabs(["🔍 개별분석", "📡 전체스캔 결과", "📸 유니버스 히스토리"])


# ═══════════════════════════════════════════════════════════════════════════════
# 탭 1: 개별분석
# ═══════════════════════════════════════════════════════════════════════════════

with tab_individual:
    RATINGS = ["-- 선택 안함 --", "AAA", "AA+", "AA", "AA-", "A+", "A", "A-", "BBB+", "BBB", "BBB-"]

    c1, c2, c3 = st.columns(3)
    inputs = []
    for i, (col, label) in enumerate(zip([c1, c2, c3], ["기업 1", "기업 2", "기업 3"]), 1):
        with col:
            st.subheader(label)
            name = st.text_input("기업명", key=f"name{i}", placeholder="예: 삼성전자")
            price = st.number_input("현재종가 (원)", key=f"price{i}", min_value=0, value=0, step=100)
            rating = st.selectbox("신용등급 (참고용)", RATINGS, key=f"rating{i}")
            if rating != "-- 선택 안함 --":
                st.caption("kisrating.com에서 해당 등급 금리 확인 후 직접 입력하세요")
            req_ret_i = st.number_input(
                "요구수익률 (%)", key=f"req_ret{i}",
                min_value=1.0, max_value=30.0, value=7.78, step=0.01, format="%.2f",
            )
            inputs.append((name, price, req_ret_i))

    clicked = st.button("🔍 계산하기", type="primary", use_container_width=True)

    if clicked:
        companies = [(n.strip(), p, r) for n, p, r in inputs if n.strip()]
        if not companies:
            st.warning("기업명을 하나 이상 입력하세요.")
        else:
            results = []
            with st.spinner("DART에서 데이터를 조회하는 중..."):
                for name, price, req_ret_i in companies:
                    try:
                        results.append(analyze(name, price, year, req_ret_i, w))
                    except Exception as e:
                        results.append({"error": str(e), "input_name": name})
            st.session_state["results"] = results
            st.session_state["params"] = {"year": year, "w": w}

    if "results" in st.session_state:
        results = st.session_state["results"]
        params = st.session_state.get("params", {"year": str(datetime.now().year - 1), "w": 0.2})

        st.divider()
        st.subheader("📊 분석 결과")

        res_cols = st.columns(len(results))
        for res, col in zip(results, res_cols):
            render_card(res, col, params["year"])

        st.divider()

        b1, b2, b3 = st.columns(3)
        valid = [r for r in results if "error" not in r]

        with b1:
            if valid:
                buf = make_excel(results, params["year"], params["w"])
                st.download_button(
                    "📥 엑셀로 내보내기",
                    data=buf,
                    file_name=f"srim_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        with b2:
            if st.button("💾 이력 저장", use_container_width=True):
                keep_keys = (
                    "company", "input_name", "current_price",
                    "roe", "fair_basic", "fair_w", "recommend_basic", "recommend_w",
                )
                entry = {
                    "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    **params,
                    "results": [{k: v for k, v in r.items() if k in keep_keys} for r in results],
                }
                save_history(entry)
                st.success("이력이 저장되었습니다.")

        with b3:
            if st.button("📋 이력 보기 / 닫기", use_container_width=True):
                st.session_state["show_history"] = not st.session_state.get("show_history", False)

    if st.session_state.get("show_history", False):
        st.divider()
        st.subheader("📋 분석 이력")
        history = load_history()
        if not history:
            st.info("저장된 이력이 없습니다.")
        else:
            rows = []
            for entry in reversed(history):
                for r in entry.get("results", []):
                    if "error" not in r:
                        rows.append({
                            "날짜": entry.get("date", ""),
                            "기준연도": entry.get("year", ""),
                            "요구수익률(%)": entry.get("req_ret", ""),
                            "실적저하율(w)": entry.get("w", ""),
                            "기업명": r.get("company", r.get("input_name", "")),
                            "현재종가": fmt(r.get("current_price")),
                            "ROE(%)": fmt(r.get("roe"), 2) if r.get("roe") is not None else "-",
                            "적정주가_기본": fmt(r.get("fair_basic")),
                            "적정주가_보수적": fmt(r.get("fair_w")),
                            "판단_기본": "✅" if r.get("recommend_basic") else "❌",
                            "판단_보수적": "✅" if r.get("recommend_w") else "❌",
                        })
            if rows:
                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
            else:
                st.info("표시할 이력이 없습니다.")


# ═══════════════════════════════════════════════════════════════════════════════
# 탭 2: 전체스캔 결과
# ═══════════════════════════════════════════════════════════════════════════════

with tab_scan:
    # 마지막 스캔 일시
    scan_data = load_scan_results()
    if scan_data:
        last_dt = scan_data[0].get("scan_dt", "알 수 없음")
        st.info(f"마지막 스캔: **{last_dt}**  |  총 {len(scan_data)}개 종목")
    else:
        st.warning("스캔 결과가 없습니다. 아래 버튼을 눌러 스캔을 시작하세요.")

    # 스캔 실행 버튼
    col_btn, col_status = st.columns([2, 8])
    with col_btn:
        run_scan = st.button("🚀 지금 스캔하기", type="primary", use_container_width=True)

    if run_scan:
        with st.spinner("스캔 중... (수백 종목 처리로 수십 분 소요됩니다)"):
            log_box = st.empty()
            log_lines = []

            def _stream_scan():
                proc = subprocess.Popen(
                    [sys.executable, str(Path(__file__).parent / "scan_all.py")],
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    bufsize=1,
                )
                for line in proc.stdout:
                    log_lines.append(line.rstrip())
                    log_box.code("\n".join(log_lines[-30:]), language=None)
                proc.wait()

            _stream_scan()

        st.success("스캔 완료! 페이지를 새로고침하면 결과가 반영됩니다.")
        scan_data = load_scan_results()

    st.divider()

    if not scan_data:
        st.stop()

    # 스코어 필드가 없는 기존 데이터 자동 재계산 (파일 덮어쓰기 없음)
    for item in scan_data:
        if "total_score" not in item:
            item.update(compute_scores(item))

    # 점수 계산 방식 설명
    with st.expander("ℹ️ 점수 계산 방식"):
        st.markdown(
            """
**종합점수 = 밸류×35% + 퀄리티×35% + 모멘텀×20% + 규모×10%**

| 팩터 | 가중치 | 기준 |
|------|--------|------|
| **밸류** | 35% | 종가/적정주가(ratio_basic) — 0.3이하→100점, 0.8이상→0점, 선형보간. low_roe 종목→0점 |
| **퀄리티** | 35% | ROE점수(5%→0, 25%→100) + 부채비율점수(200%→0, 0%→100, 없으면50) + ROE일관성(차이0→100, ≥10%→0) 평균 |
| **모멘텀** | 20% | 6개월 수익률(-20%→0, +40%→100) + 12개월 수익률(-30%→0, +60%→100) 평균. 데이터 없으면 50점(중립) |
| **규모** | 10% | 시가총액 5,000억→0점, 10조→100점 선형 |
            """
        )

    # 사이드바 필터값 읽기
    req_ret_filter = st.session_state.get("scan_req_ret", 5.0)
    ratio_max = st.session_state.get("scan_ratio_max", 0.8)
    min_cap_eok = st.session_state.get("scan_min_cap", 5000)
    sort_by = st.session_state.get("scan_sort_by", "종합점수순")

    min_cap_won = min_cap_eok * 1e8

    # 필터링
    filtered = []
    for item in scan_data:
        roe = item.get("roe", 0)
        ratio = item.get("ratio_basic")
        cap = item.get("market_cap", 0)
        low_roe = item.get("low_roe", True)

        if low_roe or roe <= req_ret_filter:
            continue
        if ratio is None or ratio >= ratio_max:
            continue
        if cap < min_cap_won:
            continue
        filtered.append(item)

    # 정렬
    if sort_by == "종합점수순":
        filtered.sort(key=lambda x: x.get("total_score", 0), reverse=True)
    else:
        filtered.sort(key=lambda x: x.get("ratio_basic") or 999)

    st.subheader(f"📋 스크리너 결과 — {len(filtered)}개 종목")
    st.caption(
        f"필터: ROE > {req_ret_filter}%  |  종가/적정주가 < {ratio_max}  |  "
        f"시가총액 ≥ {min_cap_eok:,}억원  |  정렬: {sort_by}"
    )

    if not filtered:
        st.info("조건에 맞는 종목이 없습니다. 사이드바에서 필터를 조정해보세요.")
    else:
        rows = []
        for rank, item in enumerate(filtered, 1):
            cap_eok = item["market_cap"] / 1e8
            dr = item.get("debt_ratio")
            r3 = item.get("roe_3yr_avg")
            warn = "⚠️" if item.get("cyclical_warning") else ""
            rows.append({
                "순위": rank,
                "경고": warn,
                "종목명": item["name"],
                "종합점수": f"{item['total_score']:.1f}" if item.get("total_score") is not None else "-",
                "밸류": f"{item['value_score']:.1f}" if item.get("value_score") is not None else "-",
                "퀄리티": f"{item['quality_score']:.1f}" if item.get("quality_score") is not None else "-",
                "모멘텀": f"{item['momentum_score']:.1f}" if item.get("momentum_score") is not None else "-",
                "업종": item.get("sector", ""),
                "시가총액(억원)": f"{cap_eok:,.0f}",
                "ROE(%)": f"{item['roe']:.2f}",
                "3yr ROE(%)": f"{r3:.2f}" if r3 is not None else "-",
                "부채비율(%)": f"{dr:.0f}" if dr is not None else "-",
                "적정주가_기본": f"{item['fair_basic']:,.0f}",
                "현재주가": f"{item['current_price']:,.0f}",
                "종가/적정주가": f"{item['ratio_basic']:.3f}",
            })

        df = pd.DataFrame(rows)

        def _color_score(val):
            try:
                v = float(val)
            except (ValueError, TypeError):
                return ""
            t = max(0.0, min(1.0, v / 100.0))
            r = int(220 - t * 180)
            g = int(53 + t * 114)
            b = int(69 - t * 69)
            return f"background-color: rgb({r},{g},{b}); color: white;"

        def _color_ratio(val):
            try:
                v = float(val)
            except (ValueError, TypeError):
                return ""
            lo, hi = 0.3, ratio_max
            t = max(0.0, min(1.0, (v - lo) / max(hi - lo, 0.01)))
            r = int(40 + t * 215)
            g = int(167 - t * 127)
            b = int(69 - t * 69)
            return f"background-color: rgb({r},{g},{b}); color: white;"

        def _color_warn(val):
            if val:
                return "background-color: #fff3cd; color: #856404;"
            return ""

        styled = (
            df.style
            .applymap(_color_score, subset=["종합점수", "밸류", "퀄리티", "모멘텀"])
            .applymap(_color_ratio, subset=["종가/적정주가"])
            .applymap(_color_warn, subset=["경고"])
        )
        st.dataframe(styled, use_container_width=True, hide_index=True)

        st.divider()
        export_df = pd.DataFrame([
            {
                "순위": rank,
                "종목코드": item["ticker"],
                "종목명": item["name"],
                "종합점수": item.get("total_score"),
                "밸류점수": item.get("value_score"),
                "퀄리티점수": item.get("quality_score"),
                "모멘텀점수": item.get("momentum_score"),
                "규모점수": item.get("size_score"),
                "6개월수익률(%)": item.get("mom_6m"),
                "12개월수익률(%)": item.get("mom_12m"),
                "업종": item.get("sector", ""),
                "경기민감": "⚠️" if item.get("cyclical_warning") else "",
                "시가총액(억원)": round(item["market_cap"] / 1e8, 0),
                "ROE(%)": round(item["roe"], 2),
                "3yr ROE(%)": item.get("roe_3yr_avg"),
                "부채비율(%)": item.get("debt_ratio"),
                "적정주가_기본(원)": item["fair_basic"],
                "적정주가_보수적(원)": item["fair_w"],
                "현재주가(원)": item["current_price"],
                "종가/적정주가": round(item["ratio_basic"], 4),
                "자기자본(억원)": round(item["equity"] / 1e8, 1),
                "당기순이익(억원)": round(item["net_income"] / 1e8, 1),
            }
            for rank, item in enumerate(filtered, 1)
        ])
        buf = make_scan_excel(export_df)
        st.download_button(
            "📥 엑셀로 내보내기",
            data=buf,
            file_name=f"scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ═══════════════════════════════════════════════════════════════════════════════
# 탭 3: 유니버스 히스토리
# ═══════════════════════════════════════════════════════════════════════════════

with tab_universe:
    quarters = list_snapshots()

    if not quarters:
        st.info("저장된 유니버스 스냅샷이 없습니다. 스캔을 실행하면 자동으로 저장됩니다.")
        st.stop()

    # ── 분기별 스냅샷 목록 ────────────────────────────────────────────────────
    st.subheader("📋 분기별 유니버스 현황")

    summary_rows = []
    for q in reversed(quarters):
        data = load_snapshot(q)
        if not data:
            continue
        for scan in data["scans"]:
            summary_rows.append({
                "분기": q.replace("_", " "),
                "스캔일시": scan["scan_dt"],
                "유니버스 종목수": scan["ticker_count"],
            })

    st.dataframe(pd.DataFrame(summary_rows), hide_index=True, use_container_width=True)

    st.divider()

    # ── 분기 비교 ─────────────────────────────────────────────────────────────
    st.subheader("🔍 분기 유니버스 비교")

    if len(quarters) < 2:
        st.info("분기가 2개 이상 쌓이면 비교 가능합니다. 다음 분기 스캔 후 다시 확인하세요.")
    else:
        col1, col2 = st.columns(2)
        with col1:
            q_old = st.selectbox("기준 분기 (이전)", quarters, index=0, key="q_old")
        with col2:
            q_new = st.selectbox("비교 분기 (이후)", quarters, index=len(quarters) - 1, key="q_new")

        if q_old == q_new:
            st.warning("서로 다른 분기를 선택하세요.")
        else:
            result = compare_universes(q_old, q_new)
            added     = result["added"]
            removed   = result["removed"]
            maintained = result["maintained"]

            m1, m2, m3 = st.columns(3)
            m1.metric("✅ 유지", len(maintained))
            m2.metric("🆕 신규 편입", len(added))
            m3.metric("❌ 제외 (생존편향 대상)", len(removed))

            if removed:
                st.subheader("❌ 제외된 종목 (상장폐지·조건미달)")
                st.caption("이 종목들이 분석에서 사라진 이유를 확인하세요 — 생존편향의 원인이 됩니다.")
                removed_df = pd.DataFrame(removed).rename(
                    columns={"ticker": "종목코드", "name": "종목명"}
                )
                st.dataframe(
                    removed_df.style.applymap(
                        lambda _: "background-color: #fde8e8; color: #721c24;",
                        subset=["종목명"],
                    ),
                    hide_index=True,
                    use_container_width=True,
                )

            if added:
                st.subheader("🆕 신규 편입 종목")
                added_df = pd.DataFrame(added).rename(
                    columns={"ticker": "종목코드", "name": "종목명"}
                )
                st.dataframe(
                    added_df.style.applymap(
                        lambda _: "background-color: #d4edda; color: #155724;",
                        subset=["종목명"],
                    ),
                    hide_index=True,
                    use_container_width=True,
                )
